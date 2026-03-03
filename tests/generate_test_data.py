import openpyxl
import random
import os

def generate_test_data(output_path, num_rows=100):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    def col_to_idx(col_str):
        exp = 0
        idx = 0
        for char in reversed(col_str.upper()):
            idx += (ord(char) - ord('A') + 1) * (26 ** exp)
            exp += 1
        return idx # openpyxl is 1-based for cell access

    # Mappings from import_foods.py (adjusted for 1-based openpyxl)
    cols = {
        'A': 'BLS_CODE',
        'B': 'NAME',
        'D': 'KJ',
        'G': 'KCAL',
        'M': 'PROTEIN',
        'P': 'FAT',
        'S': 'CARBS',
        'V': 'FIBRE',
        'EO': 'IRON',
        'HL': 'SUGAR',
        'LA': 'OMEGA3',
        'DO': 'VITC',
        'EF': 'MAGNESIUM',
        'ER': 'ZINC',
        'DK': 'VITB12',
        'AH': 'VITA',
        'EC': 'CALCIUM',
        'AW': 'VITD',
    }

    # Set headers (optional but good for schema)
    for col_letter, header in cols.items():
        ws[f"{col_letter}1"] = header

    for row in range(2, num_rows + 2):
        ws[f"A{row}"] = f"CODE_{row-1:05d}"
        ws[f"B{row}"] = f"Food Item {row-1}"
        ws[f"D{row}"] = random.uniform(0, 3000) # KJ
        ws[f"G{row}"] = random.uniform(0, 800)  # KCAL
        ws[f"M{row}"] = random.uniform(0, 100)  # PROTEIN
        ws[f"P{row}"] = random.uniform(0, 100)  # FAT
        ws[f"S{row}"] = random.uniform(0, 100)  # CARBS
        ws[f"V{row}"] = random.uniform(0, 50)   # FIBRE
        ws[f"EO{row}"] = random.uniform(0, 20)  # IRON
        ws[f"HL{row}"] = random.uniform(0, 100) # SUGAR
        ws[f"LA{row}"] = random.uniform(0, 5)   # OMEGA3
        ws[f"DO{row}"] = random.uniform(0, 100) # VITC
        ws[f"EF{row}"] = random.uniform(0, 500) # MAGNESIUM
        ws[f"ER{row}"] = random.uniform(0, 20)  # ZINC
        ws[f"DK{row}"] = random.uniform(0, 10)  # VITB12
        ws[f"AH{row}"] = random.uniform(0, 1000)# VITA
        ws[f"EC{row}"] = random.uniform(0, 1000)# CALCIUM
        ws[f"AW{row}"] = random.uniform(0, 20)  # VITD

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Generated {num_rows} rows at {output_path}")

if __name__ == "__main__":
    generate_test_data("/home/orchid/projects/meal-plan-analyzer-opencode/tests/data/test_foods.xlsx")
