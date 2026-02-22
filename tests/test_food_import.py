import io
import zipfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest
from django.core.management import call_command
from io import StringIO

from meals.models import Food


DATA_DIR = Path(__file__).parent / "data"
TEST_XLSX = DATA_DIR / "test_foods.xlsx"
TEST_ZIP_DATEN = DATA_DIR / "test_foods_Daten.zip"
TEST_ZIP_NO_DATEN = DATA_DIR / "test_foods_no_daten.zip"


@pytest.fixture(autouse=True)
def clear_foods():
    Food.objects.all().delete()
    yield


@pytest.mark.django_db
def test_food_import_from_xlsx():
    assert TEST_XLSX.exists(), f"Test file not found at {TEST_XLSX}"

    call_command('import_foods', str(TEST_XLSX))

    assert Food.objects.count() == 100

    wb = openpyxl.load_workbook(TEST_XLSX, data_only=True)
    ws = wb.active
    first_code = ws['A2'].value
    first_name = ws['B2'].value
    first_kcal = float(ws['G2'].value)

    food = Food.objects.get(bls_code=first_code)
    assert food.name == first_name
    assert pytest.approx(food.energy_in_kcal_per_100g) == first_kcal


@pytest.mark.django_db
def test_food_import_update_existing():
    wb = openpyxl.load_workbook(TEST_XLSX, data_only=True)
    ws = wb.active
    code = ws['A2'].value

    Food.objects.create(
        bls_code=code,
        name="Old Name",
        energy_in_kj_per_100g=0,
        energy_in_kcal_per_100g=0
    )

    call_command('import_foods', str(TEST_XLSX))

    food = Food.objects.get(bls_code=code)
    assert food.name == ws['B2'].value
    assert food.energy_in_kcal_per_100g > 0


@pytest.mark.django_db
def test_food_import_from_local_zip():
    assert TEST_ZIP_DATEN.exists(), f"Test zip not found at {TEST_ZIP_DATEN}"

    call_command('import_foods', str(TEST_ZIP_DATEN))

    assert Food.objects.count() == 100


@pytest.mark.django_db
def test_food_import_from_url():
    with open(TEST_XLSX, 'rb') as f:
        xlsx_content = f.read()

    mock_response = MagicMock()
    mock_response.read.return_value = xlsx_content
    mock_response.__enter__ = MagicMock(return_value=mock_response)
    mock_response.__exit__ = MagicMock(return_value=False)

    with patch('urllib.request.urlopen', return_value=mock_response):
        with patch('urllib.request.urlretrieve') as mock_retrieve:
            def side_effect(url, path):
                with open(path, 'wb') as f:
                    f.write(xlsx_content)
            mock_retrieve.side_effect = side_effect

            call_command('import_foods', 'https://example.com/test.xlsx')

    assert Food.objects.count() == 100


@pytest.mark.django_db
def test_food_import_from_zip_url():
    with open(TEST_ZIP_DATEN, 'rb') as f:
        zip_content = f.read()

    with patch('urllib.request.urlretrieve') as mock_retrieve:
        def side_effect(url, path):
            with open(path, 'wb') as f:
                f.write(zip_content)
        mock_retrieve.side_effect = side_effect

        call_command('import_foods', 'https://example.com/test.zip')

    assert Food.objects.count() == 100


@pytest.mark.django_db
def test_food_import_zip_no_daten_file():
    assert TEST_ZIP_NO_DATEN.exists(), f"Test zip not found at {TEST_ZIP_NO_DATEN}"

    stderr = StringIO()
    call_command('import_foods', str(TEST_ZIP_NO_DATEN), stderr=stderr)

    assert Food.objects.count() == 0
    assert "No xlsx file with" in stderr.getvalue()


@pytest.mark.django_db
def test_food_import_invalid_file_format():
    invalid_file = DATA_DIR / "invalid.txt"
    invalid_file.write_text("not an excel file")

    try:
        stderr = StringIO()
        call_command('import_foods', str(invalid_file), stderr=stderr)

        assert Food.objects.count() == 0
        assert "Invalid file format" in stderr.getvalue()
    finally:
        invalid_file.unlink()


@pytest.mark.django_db
def test_food_import_url_download_error():
    with patch('urllib.request.urlretrieve') as mock_retrieve:
        import urllib.error
        mock_retrieve.side_effect = urllib.error.URLError("Connection refused")

        stderr = StringIO()
        call_command('import_foods', 'https://example.com/test.xlsx', stderr=stderr)

        assert Food.objects.count() == 0
        assert "Failed to download" in stderr.getvalue()


@pytest.mark.django_db
def test_food_import_file_not_found():
    stderr = StringIO()
    call_command('import_foods', '/nonexistent/path/file.xlsx', stderr=stderr)

    assert Food.objects.count() == 0
    assert "File not found" in stderr.getvalue()
